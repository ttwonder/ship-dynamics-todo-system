"""Build the browser-safe blank Itinerary workbook template.

The source workbook is supplied by the product owner. This script keeps the
formula-version sheet's layout, styles, merged two-row records, validation,
and A:M print layout while removing all sample business values.
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


def build(source: Path, output: Path) -> dict[str, object]:
    workbook = load_workbook(source)
    if "有公式版本" not in workbook.sheetnames:
        raise ValueError("source workbook is missing 有公式版本")

    worksheet = workbook["有公式版本"]
    for candidate in list(workbook.worksheets):
        if candidate is not worksheet:
            workbook.remove(candidate)
    worksheet.title = "Itinerary Template"

    cleared = 0
    for row in worksheet.iter_rows(min_row=4, max_row=worksheet.max_row, min_col=1, max_col=23):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.value is not None:
                cleared += 1
            cell.value = None
            cell.comment = None
            cell.hyperlink = None

    worksheet.print_area = f"A1:M{worksheet.max_row}"
    worksheet.sheet_view.showGridLines = False
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return {
        "source": str(source),
        "output": str(output),
        "sheet": worksheet.title,
        "rows": worksheet.max_row,
        "columns": worksheet.max_column,
        "cleared_cells": cleared,
        "print_area": worksheet.print_area,
        "merged_ranges": len(worksheet.merged_cells.ranges),
        "data_validations": len(worksheet.data_validations.dataValidation),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    print(json.dumps(build(args.source, args.output), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
